# read the content of several files
# 1 line per row 
# 1 file per column
import openpyxl

wb = openpyxl.Workbook()
ws = wb.active

with open('text1.txt') as file:
    text1 = file.readlines()
with open('text2.txt') as file:
    text2 = file.readlines()

# Split sentences by comma
for group in text1:
    lines_text1 = group.split('.')
for group in text2:
    lines_text2 = group.split('.')


for index, sentence in enumerate(lines_text1, 1):
    cell_to_add_text = ws.cell(row=index, column=1)
    cell_to_add_text.value = sentence

for index, sentence in enumerate(lines_text2, 1):
    cell_to_add_text = ws.cell(row=index, column=2)
    cell_to_add_text.value = sentence

wb.save("text to spreadsheet.xlsx")
