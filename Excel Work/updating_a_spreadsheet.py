#Change the price of 3 items in a spreadsheet
#celery = 1.19
#Garlic = 3.07
#Lemon = 1.27
import openpyxl

# Open the spreadsheet file.
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

#Store the produce types and their updated prices
price_updates = {'Garlic':3.07, 
                 'Celery': 1.19,
                 'Lemon': 1.27}

# For each row, check whether the value in column A is Celery, Garlic,
# or Lemon. If it is, update the price in column B.

for row_num in range(2, sheet.max_row): #skip the 1st rox
    produce_name = sheet.cell(row=row_num, column = 1).value
    if produce_name in price_updates:
        sheet.cell(row=row_num, column=2).value = price_updates[produce_name]

#Save the spreadsheet to a new file (so that you don’t lose the old spreadsheet,
# just in case).
wb.save('updated_produce_sales.xlsx')
