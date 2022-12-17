# read the content a spreeadsheet and write new files
# 1 column = 1 file
# 1 row = 1 sentence

import openpyxl

wb = openpyxl.load_workbook('text to spreadsheet.xlsx')
ws = wb.active

for col in range(1, ws.max_column+1):
    f = open(f"spreadsheet to text{col}.txt", "x")
    for row in range(1, ws.max_row+1):
        cell_value = ws.cell(row=row, column=col).value
        if cell_value:
            f.write(str(cell_value) + '.')
    f.close()
