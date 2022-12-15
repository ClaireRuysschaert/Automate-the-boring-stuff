# Takes a number N from the command line and creates an N×N multiplication table in an Excel spreadsheet

import sys

import openpyxl
from openpyxl.styles import Font

wb = openpyxl.Workbook() #create a blank
ws = wb['Sheet'] #get the worksheet object

N = int(sys.argv[1])

# Numéroter la colonne A et la ligne 1 de 1 à N
for i in range(2, N+2):
    ws.cell(row=1, column = i).value = i-1
    ws.cell(row=i, column = 1).value = i-1


# Multiplier les cellules B1 et A2 et mettre le résultat dans la cellule B2
#ws['B2'] = '=B1*A2'
for each_row in range(2,N+2):
    for i in range(2,N+2):
        product1 = ws.cell(row=1, column=i).value 
        product2 = ws.cell(row=each_row, column=1).value
        ws.cell(row=each_row, column=i).value = product1 * product2

# Font the first row and column
bold_font = Font(bold = True)
for i in range(1, 101):
    ws.cell(row=i, column=1).font = bold_font
    ws.cell(row=1, column=i).font = bold_font

wb.save('Multiplication Table.xlsx')
