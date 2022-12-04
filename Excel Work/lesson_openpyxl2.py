import openpyxl
from openpyxl.styles import Font

wb = openpyxl.Workbook() #create a blank
sheet = wb.active

#Create and delete sheets
wb.create_sheet(index=1, title = 'YUKI')
sheet.title = 'TDB'
wb.create_sheet(index=2, title='ANN')
del wb['ANN']

#Font sheets
sheet['A1'] = 'suki'
sheet_font = Font(name = 'Calibri', size=24, italic=True)
sheet["A1"].font = sheet_font

#Formulas
sheet['B1'] = 200
sheet['B2'] = 300
sheet['B3'] = '=SUM(B1:B2)'

#Row height and column width
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20

#Frozen panes
sheet.freeze_panes = 'A2' #row 1 frozen
# B1 = columnA ; C1 = column A + B ; C2 = row1 + column A + B
# A1 = no frozen panes

wb.save('TDB.xlsx')
