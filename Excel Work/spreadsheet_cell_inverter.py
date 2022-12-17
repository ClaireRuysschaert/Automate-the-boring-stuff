import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
ws = wb.active
new_wb = openpyxl.Workbook()
new_ws = new_wb.active

for row in range(1, ws.max_row +1):
    for col in range(1, ws.max_column+1):
        new_ws.cell(col, row).value = ws.cell(row, col).value

new_wb.save('inverted file.xlsx')
