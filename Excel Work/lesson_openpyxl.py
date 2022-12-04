import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('example.xlsx')

#Getting cells from the sheets
sheet = wb['Sheet1']
for i in range(1, 8, 1):
    print(i, sheet.cell(row=i, column=2).value, sheet.cell(row=i, column=3).value)

#Converting Between Column Letters and Numbers
print(get_column_letter(sheet.max_column), column_index_from_string('A'))

#Getting Rows and Columns from the Sheets
##Get all cells from A1 to C3 = 1 tuple for each row

for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
       print(cellObj.coordinate, cellObj)
