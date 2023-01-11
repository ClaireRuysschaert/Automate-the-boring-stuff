# sendDuesReminders.py - Sends emails based on payment status in spreadsheet.

import os
import smtplib

import openpyxl
from dotenv import load_dotenv

# Open the spreadsheet and get the latest dues status.
wb = openpyxl.load_workbook("duesRecords.xlsx")
sheet = wb.get_sheet_by_name("Sheet1")
lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value

# Check each member's payment status.
unpaidMembers = {}
for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != "paid":
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email

# Log in to email account.
smtpObj = smtplib.SMTP("smtp.example.com", 587)
smtpObj.ehlo()
smtpObj.starttls()
load_dotenv()
smtp_pw = os.getenv("smtp_password")
smtpObj.login("yukimargoulin@gmail.com", smtp_pw)

# Send out reminder emails.
for name, email in unpaidMembers.items():
    body = f"Subject: {latestMonth} dues unpaid.\nDear {name},\nRecords show"
    " that you have not paid dues for {latestMonth}. Please make this payment"
    " as soon as possible. Thank you!'"
    print(f"Sending email to {email}...")
    sendmailStatus = smtpObj.sendmail("yukimargoulin@gmail.com", email, body)
    if sendmailStatus != {}:
        print(f"There was a problem sending email to {email}:{sendmailStatus}")
smtpObj.quit()
